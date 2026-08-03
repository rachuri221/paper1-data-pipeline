"""
analysis/10_build_supplementary_report.py

Assembles the SCF benchmarking tables and NFWBS regression results into a
second workbook, kept separate from the NFCS report
(outputs/models/NFCS_Financial_Ed_Full_Report.xlsx) because these two
datasets play a fundamentally different role in the project:

  - SCF: aggregate national benchmarks, not respondent microdata, not
    pooled into any regression.
  - NFWBS: respondent microdata, but with no predictor equivalent to
    NFCS's M20, modeling a related-but-distinct construct (informal
    parental financial socialization, not formal financial education).

Keeping them in a separate report file mirrors that they answer different
questions and should not be cited interchangeably with the primary NFCS
regression results.

Output: outputs/models/SCF_and_NFWBS_Supplementary_Report.xlsx
"""

import os
import sys
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import utils

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13)
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="444444")
CELL_FONT = Font(name="Arial", size=10)
BOLD_CELL = Font(name="Arial", size=10, bold=True)
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SIG_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

TABLE_START_ROW = 4


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")


def write_table_sheet(writer, df, sheet_name, title, subtitle, sig_col=None, col_widths=None):
    df.to_excel(writer, sheet_name=sheet_name, startrow=TABLE_START_ROW, index=False)
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT

    header_row = TABLE_START_ROW + 1
    ncols = len(df.columns)
    style_header(ws, header_row, ncols)

    numeric_3dp = {"Coefficient", "SE", "CI_low", "CI_high", "Value", "Median_2019", "Median_2022",
                   "Mean_2019", "Mean_2022", "y2010", "y2013", "y2016", "y2019", "y2022"}
    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + i
        for c, colname in enumerate(df.columns, 1):
            cell = ws.cell(row=excel_row, column=c)
            cell.font = CELL_FONT
            if colname in numeric_3dp:
                cell.number_format = "0.00"
            if colname == "p-value":
                cell.number_format = "0.0000"
        if sig_col and sig_col in df.columns and row[sig_col] == "Yes":
            for c in range(1, ncols + 1):
                ws.cell(row=excel_row, column=c).fill = SIG_FILL

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


METHODOLOGY_TEXT = [
    "", "SCF (SURVEY OF CONSUMER FINANCES) -- BENCHMARKING ROLE",
    "Source: Federal Reserve 'Changes in U.S. Family Finances from 2019 to 2022' bulletin",
    "(scf23.pdf). Tables extracted programmatically (analysis/07_scf_extraction.py) via",
    "pdfplumber + regex parsing of the bulletin's text tables -- see docs/cleaning_log.md",
    "for extraction details and page sources.",
    "",
    "SCF publishes only AGGREGATE summary statistics (medians, means, percentages by",
    "characteristic group), not respondent-level microdata. It therefore cannot enter any",
    "regression model in this project. Its role is strictly descriptive benchmarking: how",
    "does the NFCS young-adult sample's financial position compare to national context?",
    "The SCF_vs_NFCS_Comparison sheet is a DESCRIPTIVE JUXTAPOSITION ONLY -- SCF and NFCS",
    "differ in population, survey year, and question wording, so this is not a statistical",
    "comparison and no significance test is implied.",
    "",
    "NFWBS (NATIONAL FINANCIAL WELL-BEING SURVEY) -- SCOPE AND WHY IT'S NOT POOLED",
    "Source: CFPB National Financial Well-Being Survey PUF, 2016, N=6,394",
    "(NFWBS_PUF_2016_data.csv). Real respondent-level microdata.",
    "",
    "NFWBS has NO item equivalent to NFCS's M20 (formal financial-education exposure --",
    "'was financial education offered by a school/employer, did you participate'). It",
    "therefore CANNOT support the same predictor as the NFCS models and is never pooled",
    "with them under any circumstance.",
    "",
    "What NFWBS does have is FINSOC2_1 through FINSOC2_7: seven items on whether a parent",
    "engaged in specific financial-socialization behaviors (discussed money, taught",
    "budgeting, gave an allowance, provided a savings account, etc.) while the respondent",
    "was growing up. This script uses a 0-7 sum of those items as a predictor of INFORMAL",
    "financial socialization -- a related but conceptually DISTINCT construct from formal",
    "financial education. Every output in this workbook labels it as such; treat any",
    "similarity to the NFCS findings as suggestive triangulation across a related construct,",
    "not as replication of the same effect.",
    "",
    "Age proxy: NFWBS's 'agecat' variable's youngest bracket (agecat==1, n=414) is used the",
    "same way NFCS's 18-24 bucket is used -- as a disclosed, imperfectly-bounded proxy for",
    "'youngest adults in the sample,' not a precisely-defined age range. The exact upper",
    "bound of agecat==1 was not independently confirmed from the codebook text available at",
    "pipeline build time (the codebook confirms categories 6-8 correspond to the survey's",
    "age-62+ oversample, which anchors the top end of the scale but not category 1's cutoff).",
    "",
    "*** KEY FINDING ***",
    "Parental financial socialization (0-7 scale) is significantly associated, in the",
    "expected direction, with ALL FOUR outcomes modeled: higher financial well-being",
    "(FWBscore, p<.0001), higher odds of a savings habit (OR=1.19, p<.0001), higher odds of",
    "paying credit cards in full (OR=1.15, p<.0001), and LOWER odds of difficulty making",
    "ends meet (OR=0.91, p<.0001). This is a clean, consistent pattern -- but remember it is",
    "informal parental socialization, not formal financial-education programming, and",
    "cannot be cited as evidence for or against the NFCS-based paper's hypothesis about",
    "school/employer-based financial education specifically.",
    "",
    "CAUSALITY CAVEAT (same structure as the NFCS analysis)",
    "NFWBS is cross-sectional: parental socialization is recalled retrospectively by adult",
    "respondents, and current financial well-being is measured at the same interview. No",
    "random assignment, no panel structure, and recall of childhood socialization may itself",
    "be correlated with current financial well-being (e.g., people doing better financially",
    "may be more likely to recall their parents positively). Correlational, not causal.",
]


def build_methodology_sheet(writer):
    ws = writer.book.create_sheet("Methodology", 0)
    ws.cell(row=1, column=1, value="Supplementary Datasets — SCF Benchmarking & NFWBS Well-Being Analysis").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Prepared by Lydia (Statistician)").font = SUBTITLE_FONT
    r = 4
    for line in METHODOLOGY_TEXT:
        cell = ws.cell(row=r, column=1, value=line)
        if line.startswith("***"):
            cell.font = Font(name="Arial", bold=True, size=10, color="990000")
            cell.fill = FLAG_FILL
        elif line.isupper() and line.strip():
            cell.font = BOLD_CELL
        else:
            cell.font = CELL_FONT
        r += 1
    ws.column_dimensions["A"].width = 100


if __name__ == "__main__":
    T = utils.TABLES_DIR
    out_path = os.path.join(utils.MODELS_DIR, "SCF_and_NFWBS_Supplementary_Report.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="_placeholder")

        scf_summary = pd.read_csv(os.path.join(T, "scf_benchmark_summary.csv"))
        write_table_sheet(writer, scf_summary, "SCF_Benchmark_Summary",
                           "SCF Benchmark Summary — National Context (2022 unless noted)",
                           "Source: Federal Reserve SCF Bulletin, extracted programmatically — aggregate data, not microdata",
                           col_widths=[70, 12, 14])

        scf_t1 = pd.read_csv(os.path.join(utils.CLEANED_DIR, "scf_table1_income.csv"))
        write_table_sheet(writer, scf_t1, "SCF_Table1_Income",
                           "SCF Table 1 — Income by Characteristic (extracted from scf23.pdf)",
                           "Thousands of 2022 dollars — full extracted table, all 32 rows",
                           col_widths=[38, 12, 12, 14, 12, 12, 14])

        scf_t2 = pd.read_csv(os.path.join(utils.CLEANED_DIR, "scf_table2_networth.csv"))
        write_table_sheet(writer, scf_t2, "SCF_Table2_NetWorth",
                           "SCF Table 2 — Net Worth by Characteristic (extracted from scf23.pdf)",
                           "Thousands of 2022 dollars — full extracted table, all 31 rows",
                           col_widths=[38, 12, 12, 14, 12, 12, 14])

        scf_t5 = pd.read_csv(os.path.join(utils.CLEANED_DIR, "scf_table5_debt_burden.csv"))
        write_table_sheet(writer, scf_t5, "SCF_Table5_DebtBurden",
                           "SCF Table 5 — Debt Burden & Credit Market Experiences (extracted from scf23.pdf)",
                           "% of all families unless noted — full extracted table, all 16 rows",
                           col_widths=[55, 10, 10, 10, 10, 10])

        comparison_path = os.path.join(T, "scf_vs_nfcs_comparison.csv")
        if os.path.exists(comparison_path):
            comparison_df = pd.read_csv(comparison_path)
            if "Note" in comparison_df.columns and len(comparison_df.columns) == 1:
                pass  # NFCS file wasn't available when this was generated; skip gracefully
            else:
                write_table_sheet(writer, comparison_df, "SCF_vs_NFCS_Comparison",
                                   "SCF vs. NFCS — Descriptive Juxtaposition (NOT a statistical comparison)",
                                   "Different surveys, populations, years, and question wording — see Methodology",
                                   col_widths=[35, 45, 25, 70])

        nfwbs_desc = pd.read_csv(os.path.join(T, "nfwbs_descriptive_stats.csv"))
        write_table_sheet(writer, nfwbs_desc, "NFWBS_Descriptive_Stats",
                           "NFWBS Descriptive Statistics (N=6,394)",
                           "",
                           col_widths=[55, 14])

        nfwbs_results = pd.read_csv(os.path.join(T, "nfwbs_regression_results.csv"))
        write_table_sheet(writer, nfwbs_results, "NFWBS_Regression_Results",
                           "NFWBS — Parental Financial Socialization -> Well-Being Outcomes",
                           "Informal socialization (FINSOC2 items), NOT formal financial education — see Methodology",
                           sig_col="Sig (p<.05)", col_widths=[45, 20, 8, 14, 10, 10, 10, 12, 12, 55])

        build_methodology_sheet(writer)
        wb = writer.book
        wb.move_sheet("Methodology", offset=-(len(wb.sheetnames) - 1))
        del wb["_placeholder"]

    print(f"Saved: {out_path}")
