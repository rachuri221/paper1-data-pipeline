"""
analysis/05_build_final_report.py

Assembles every table produced by scripts 02-04 into one formatted Excel
workbook. Uses pandas.ExcelWriter (openpyxl engine) to write each table via
DataFrame.to_excel() rather than hand-indexing cells/rows — this avoids an
entire class of off-by-one / infinite-loop bugs that manual openpyxl row
bookkeeping is prone to (one such bug was caught and fixed during testing
of an earlier version of this pipeline; see git history).

Sheets:
  1. Methodology            - data sources, harmonization decisions, caveats
  2. Descriptive_Stats       - pooled sample descriptives
  3. Primary_2024            - full 2024 sample model
  4. Subgroup_2024_18to24    - 2024 young-adult subgroup, Bonferroni-corrected
  5. ChiSquare_CrossCheck    - independent bivariate cross-validation (pingouin)
  6. Pooled_AllWaves         - pooled 2018+2021+2024 full-sample model
  7. Pooled_Subgroup_18to24  - pooled young-adult subgroup, Bonferroni-corrected
  8. ByWave_Robustness       - same spec run separately per wave

Output: outputs/models/NFCS_Financial_Ed_Full_Report.xlsx
"""

import os
import sys
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import utils

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13)
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="444444")
CELL_FONT = Font(name="Arial", size=10)
BOLD_CELL = Font(name="Arial", size=10, bold=True)
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SIG_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
WARN_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

TITLE_ROW = 1
SUBTITLE_ROW = 2
TABLE_START_ROW = 4  # 0-indexed for to_excel's startrow; header lands on row 5 in Excel (1-indexed)


def write_table_sheet(writer, df: pd.DataFrame, sheet_name: str, title: str, subtitle: str,
                       sig_col: str | None = None, col_widths: list | None = None):
    df.to_excel(writer, sheet_name=sheet_name, startrow=TABLE_START_ROW, index=False)
    ws: Worksheet = writer.sheets[sheet_name]

    ws.cell(row=TITLE_ROW, column=1, value=title).font = TITLE_FONT
    ws.cell(row=SUBTITLE_ROW, column=1, value=subtitle).font = SUBTITLE_FONT

    header_excel_row = TABLE_START_ROW + 1  # to_excel startrow is 0-indexed, openpyxl rows are 1-indexed
    ncols = len(df.columns)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_excel_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    numeric_fmt_3dp = {"OR", "OR_CI_low", "OR_CI_high", "Cohen's d (approx)", "chi2", "Cramer's V"}
    numeric_fmt_4dp = {"p-value"}
    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = header_excel_row + 1 + i
        for c, colname in enumerate(df.columns, 1):
            cell = ws.cell(row=excel_row, column=c)
            cell.font = CELL_FONT
            if colname in numeric_fmt_3dp:
                cell.number_format = "0.000"
            if colname in numeric_fmt_4dp:
                cell.number_format = "0.0000"
        if sig_col and sig_col in df.columns and row[sig_col] == "Yes":
            for c in range(1, ncols + 1):
                ws.cell(row=excel_row, column=c).fill = SIG_FILL

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_excel_row + 1, column=1).coordinate


def highlight_reversed_direction(writer, df: pd.DataFrame, sheet_name: str):
    """For the ByWave_Robustness sheet: overwrite green with red where a
    significant result runs opposite the paper's hypothesized direction
    (more financial-ed exposure associated with WORSE overspending/overdraft
    behavior, rather than better)."""
    ws: Worksheet = writer.sheets[sheet_name]
    header_excel_row = TABLE_START_ROW + 1
    ncols = len(df.columns)
    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = header_excel_row + 1 + i
        if row["Sig (p<.05)"] == "Yes":
            reversed_direction = (
                row["Outcome"] in ["Spends more than income (past year)", "Overdraws checking account"]
                and row["OR"] > 1
            )
            if reversed_direction:
                for c in range(1, ncols + 1):
                    ws.cell(row=excel_row, column=c).fill = WARN_FILL
    note_row = header_excel_row + len(df) + 2
    ws.cell(row=note_row, column=1,
            value="Red = significant, OPPOSITE the hypothesized direction. Green = significant, expected direction.")
    ws.cell(row=note_row, column=1).font = Font(name="Arial", italic=True, size=9, bold=True, color="990000")


METHODOLOGY_TEXT = [
    "", "DATA",
    "NFCS State-by-State Survey, three waves: 2018, 2021, 2024 (single-wave and",
    "pooled models both reported). Investor Survey (Inv) files are EXCLUDED from",
    "all models: that sample is screened to investment-account holders and is",
    "not representative of the target 'underserved young adult' population.",
    "CFPB National Financial Well-Being Survey PUF (2016) was evaluated for",
    "pooling but excluded: no item comparable to NFCS's M20 predictor exists.",
    "",
    "HARMONIZATION ACROSS WAVES (read before citing pooled coefficients)",
    "- Gender (A50A) does not exist in the 2018 questionnaire -> dropped from",
    "  ALL pooled and by-wave models for cross-wave consistency. It IS included",
    "  in the single-wave 2024 model (Primary_2024 / Subgroup_2024_18to24 sheets),",
    "  so those numbers are not directly comparable to the pooled sheets.",
    "- Income: A8 in 2018 (8 categories) vs A8_2021 in 2021/2024 (10 categories,",
    "  finer top-end splits). Treated as an ordinal control only.",
    "- Age bucketing (A3Ar_w) is consistent across all three waves: 18-24 is the",
    "  youngest available bucket, used as a disclosed proxy for the paper's",
    "  18-22 target population. NFCS does not sample anyone under 18.",
    "- Wave fixed effects (C(wave)) included in pooled models to absorb level",
    "  shifts across survey years unrelated to the effect of interest.",
    "",
    "CROSS-VALIDATION",
    "Each 2024 primary-model association is checked against an independent",
    "chi-square test of unadjusted bivariate association (pingouin), see the",
    "ChiSquare_CrossCheck sheet. Two outcomes ('Spends more than income' and",
    "'Carries interest-bearing CC balance') are significant unadjusted but NOT",
    "significant after controls -- meaning the raw association is explained",
    "by demographic composition (income, age, education), not by financial-ed",
    "exposure itself. This is a meaningful distinction for the paper's framing.",
    "",
    "*** KEY FINDING REQUIRING CAREFUL FRAMING ***",
    "Financial-education exposure is associated with HIGHER odds of",
    "overspending and overdrafting in the pooled model (both p<.001) --",
    "opposite the paper's hypothesis. Driven mainly by 2018/2021; 2024 alone",
    "was null on both. Consistent with a reverse-causality story (financially-",
    "struggling people more likely to seek/be directed to education). This",
    "should be reported in the paper, not omitted.",
    "",
    "*** ROBUSTNESS GOOD NEWS ***",
    "'Has emergency fund' is significant, same direction, in all three waves",
    "individually, the full pooled model, the 2024-only model, AND the pooled",
    "18-24 subgroup even after Bonferroni correction. Strongest, most",
    "defensible finding in this analysis.",
    "",
    "*** STANDING CAUSALITY CAVEAT ***",
    "NFCS waves are independent repeated cross-sections, not a panel; no",
    "individual is followed over time. Financial-education exposure and",
    "current behavior are measured in the same interview -- no temporal",
    "ordering, no random assignment. Multi-wave replication increases",
    "confidence the ASSOCIATION is real and stable, but does not establish",
    "causation. See docs/analytical_decisions.md for the full discussion.",
    "",
    "SIGNIFICANCE THRESHOLDS: Primary/pooled analyses p<.05. Subgroup analyses",
    "Bonferroni-corrected across 5 outcome tests, alpha = 0.01/5 = .002",
    "(per lab statistical protocol).",
]


def build_methodology_sheet(writer):
    ws = writer.book.create_sheet("Methodology", 0)
    ws.cell(row=1, column=1, value="Paper 1 — Financial Literacy Education and Youth Behavior").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Statistical Analysis Documentation — Lydia (Statistician)").font = SUBTITLE_FONT
    r = 4
    for line in METHODOLOGY_TEXT:
        cell = ws.cell(row=r, column=1, value=line)
        if line.startswith("***"):
            cell.font = Font(name="Arial", bold=True, size=10, color="990000")
            cell.fill = FLAG_FILL
        elif line.isupper() and line.strip():
            cell.font = BOLD_CELL
        else:
            cell.font = CELL_FONT
        r += 1
    ws.column_dimensions["A"].width = 100


if __name__ == "__main__":
    T = utils.TABLES_DIR
    out_path = os.path.join(utils.MODELS_DIR, "NFCS_Financial_Ed_Full_Report.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # placeholder sheet removed after methodology is inserted at index 0
        pd.DataFrame().to_excel(writer, sheet_name="_placeholder")

        desc_df = pd.read_csv(os.path.join(T, "descriptive_stats.csv"))
        write_table_sheet(writer, desc_df, "Descriptive_Stats",
                           "Descriptive Statistics — Pooled 2018+2021+2024",
                           "",
                           col_widths=[55, 12, 22])

        primary_df = pd.read_csv(os.path.join(T, "2024_primary_results.csv"))
        write_table_sheet(writer, primary_df, "Primary_2024",
                           "Primary Model — 2024 Full Sample",
                           "Threshold p<.05 | Controls: gender, race, education, income, age group",
                           sig_col="Sig (p<.05)", col_widths=[38, 8, 12, 10, 10, 12, 12, 16, 10, 12])

        subgroup_df = pd.read_csv(os.path.join(T, "2024_subgroup_results.csv"))
        bonf_col_24 = [c for c in subgroup_df.columns if "Bonferroni" in c][0]
        write_table_sheet(writer, subgroup_df, "Subgroup_2024_18to24",
                           "Subgroup Model — 2024, Ages 18-24",
                           "Bonferroni-corrected across 5 tests: alpha=.002",
                           sig_col=bonf_col_24, col_widths=[38, 8, 10, 12, 12, 16, 10, 22, 22])

        chi_df = pd.read_csv(os.path.join(T, "2024_chi_square_crossvalidation.csv"))
        write_table_sheet(writer, chi_df, "ChiSquare_CrossCheck",
                           "Independent Cross-Validation — Chi-Square Test (unadjusted, pingouin)",
                           "Bivariate association only, no controls — compare against Primary_2024 for confounding checks",
                           sig_col="Sig (p<.05)", col_widths=[38, 8, 12, 8, 12, 12, 12])

        pooled_df = pd.read_csv(os.path.join(T, "pooled_results.csv"))
        write_table_sheet(writer, pooled_df, "Pooled_AllWaves",
                           "Pooled Model — 2018+2021+2024 Combined, Wave Fixed Effects",
                           "Threshold p<.05 | Controls: race, education, income, age group, wave FE (gender excluded)",
                           sig_col="Sig (p<.05)", col_widths=[38, 8, 10, 12, 12, 16, 10, 12])

        pooled_sub_df = pd.read_csv(os.path.join(T, "pooled_subgroup_results.csv"))
        bonf_col_pool = [c for c in pooled_sub_df.columns if "Bonferroni" in c][0]
        write_table_sheet(writer, pooled_sub_df, "Pooled_Subgroup_18to24",
                           "Pooled Young-Adult Subgroup (18-24) — 2018+2021+2024 Combined",
                           "Bonferroni-corrected across 5 tests: alpha=.002 | Controls: race, education, income, wave FE",
                           sig_col=bonf_col_pool, col_widths=[38, 8, 10, 12, 12, 16, 10, 22, 22])

        wave_df = pd.read_csv(os.path.join(T, "by_wave_results.csv"))
        write_table_sheet(writer, wave_df, "ByWave_Robustness",
                           "By-Wave Robustness — Same Spec Run Separately in Each Wave",
                           "Checks whether pooled results reflect a stable pattern or are driven by a single wave",
                           sig_col="Sig (p<.05)", col_widths=[8, 38, 8, 10, 12, 12, 10, 12])
        highlight_reversed_direction(writer, wave_df, "ByWave_Robustness")

        build_methodology_sheet(writer)
        # move Methodology to front, drop placeholder
        wb = writer.book
        wb.move_sheet("Methodology", offset=-(len(wb.sheetnames) - 1))
        del wb["_placeholder"]

    print(f"Saved: {out_path}")
