"""
analysis/12_build_knowledge_score_report.py

Assembles the knowledge-score model results (analysis/11) into a formatted
Excel workbook, same visual format as the prior NFCS/SCF/NFWBS reports.

Output: outputs/models/Knowledge_Score_Underserved_18to24_Report.xlsx
"""

import os
import sys
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import utils

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13)
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="444444")
CELL_FONT = Font(name="Arial", size=10)
BOLD_CELL = Font(name="Arial", size=10, bold=True)
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SIG_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
WARN_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

TABLE_START_ROW = 4


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")


def write_results_sheet(writer, df: pd.DataFrame):
    df.to_excel(writer, sheet_name="Results", startrow=TABLE_START_ROW, index=False)
    ws = writer.sheets["Results"]
    ws.cell(row=1, column=1, value="Financial Knowledge Score -> Behavioral Outcomes, 18-24").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="NFCS 2018 & 2021 (separate models, not pooled) | OR per +1 point on 0-5 knowledge score").font = SUBTITLE_FONT

    header_row = TABLE_START_ROW + 1
    ncols = len(df.columns)
    style_header(ws, header_row, ncols)

    bonf_col = [c for c in df.columns if "Bonferroni" in c][0]
    numeric_cols = {"OR (per point)", "OR_CI_low", "OR_CI_high",
                     "Knowledge score mean (this panel)", "Knowledge score SD (this panel)"}
    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + i
        for c, colname in enumerate(df.columns, 1):
            cell = ws.cell(row=excel_row, column=c)
            cell.font = CELL_FONT
            if colname in numeric_cols:
                cell.number_format = "0.000"
            if colname == "p-value":
                cell.number_format = "0.0000"
        if row[bonf_col] == "Yes":
            for c in range(1, ncols + 1):
                ws.cell(row=excel_row, column=c).fill = SIG_FILL
        elif row["Outcome"] == "Spends more than income (past year)" and row.get("OR (per point)", 1) and \
                pd.notna(row.get("OR (per point)")) and row["OR (per point)"] > 1 and row[bonf_col] == "Yes":
            for c in range(1, ncols + 1):
                ws.cell(row=excel_row, column=c).fill = WARN_FILL

    widths = [8, 26, 38, 8, 14, 12, 12, 12, 22, 22, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


METHODOLOGY_TEXT = [
    "", "PREDICTOR",
    "Financial knowledge score, NFCS 'Big Five' quiz (M6-M10): compound interest,",
    "inflation, bond prices, mortgage true/false, diversification true/false. 0-5 points,",
    "one per correct answer. 'Don't know' scored as incorrect (0); 'Refused' treated as",
    "missing for that item -- if ANY item is refused, the respondent's total score is set",
    "to missing rather than computed from fewer than 5 items. This is NOT the same",
    "predictor as the M20-based models (financial-education EXPOSURE) -- this measures",
    "financial knowledge/literacy, a related but distinct construct.",
    "",
    "OUTCOMES (same 4 as the M20-based models, binary)",
    "has_emergency_fund (J5), pays_cc_full (F2_1), overspends (J3), overdraws_checking (B4).",
    "",
    "DATASETS",
    "NFCS State-by-State, 2018 and 2021 -- run SEPARATELY, not pooled (matches the two",
    "waves used in the paper). Gender (A50A) is not collected in the 2018 questionnaire,",
    "so it is excluded as a covariate from the 2018 models and included in the 2021",
    "models -- same constraint as the M20-based pooled models (see",
    "docs/analytical_decisions.md Sec 4), applied here per-wave since these are not pooled.",
    "",
    "*** SAMPLE DEFINITION: WHY INCOME-BAND CUTOFF, NOT FPL PERCENTAGE ***",
    "NFCS has no household-size variable and only banded (not continuous) income, so an",
    "exact percent-of-Federal-Poverty-Level cannot be computed. 'Underserved' is defined",
    "as household income under $35,000/year (NFCS income bands 1-3, identical cutoffs in",
    "both waves), NOT a literal FPL percentage. An approximate FPL proxy (household size",
    "assembled from living-arrangement + dependent-children items) was considered and",
    "rejected: that proxy is least reliable exactly for 18-24-year-olds living with",
    "parents or roommates -- likely a large share of this sample -- where 'household",
    "income' as reported is ambiguous (is it the respondent's own income, or the",
    "household they live in). The income-band cutoff avoids introducing that bias, at",
    "the cost of being a blunter definition of 'underserved' than a true FPL percentage",
    "would be. See docs/analytical_decisions.md Sec 15 for the full reasoning.",
    "",
    "CORRECTION",
    "Bonferroni across the 4 outcome tests WITHIN each wave x sample panel (4 panels",
    "total: 2018 underserved, 2018 all-income, 2021 underserved, 2021 all-income).",
    "alpha = 0.01 / 4 = .0025 per panel.",
    "",
    "EFFECT SIZE / UNITS",
    "Odds ratio per +1 POINT on the 0-5 knowledge scale (not per SD). Each panel's",
    "knowledge-score mean and SD are reported alongside every result row so a reviewer",
    "can convert to a per-SD effect if needed: OR_per_SD = OR_per_point ^ SD.",
    "",
    "*** KEY PATTERN ***",
    "2018 is strongly and consistently significant across nearly all outcomes and both",
    "samples, all in the expected direction (higher knowledge -> more emergency fund,",
    "more full CC payment, less overspending, less overdraft). 2021 is much weaker: the",
    "underserved (<$35k) 18-24 sample shows NOTHING significant even at uncorrected",
    "p<.05 in 2021, and the 2021 all-income sample only holds for 2 of 4 outcomes after",
    "Bonferroni correction. This wave-to-wave divergence is real (not a coding error --",
    "checked against raw value distributions before running) and should be discussed in",
    "the paper rather than only reporting the stronger 2018 numbers. Possible",
    "explanations include COVID-era disruption to the 2021 sample's financial behavior",
    "patterns, or reduced power in the smaller 2021 underserved subgroup (N=1,538-1,723",
    "across outcomes vs. 2018's N=792-1,647) -- these are hypotheses, not confirmed.",
]


def build_methodology_sheet(writer):
    ws = writer.book.create_sheet("Methodology", 0)
    ws.cell(row=1, column=1, value="Financial Knowledge Score Models — Underserved 18-24").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Prepared by Lydia (Statistician)").font = SUBTITLE_FONT
    r = 4
    for line in METHODOLOGY_TEXT:
        cell = ws.cell(row=r, column=1, value=line)
        if line.startswith("***"):
            cell.font = Font(name="Arial", bold=True, size=10, color="990000")
            cell.fill = FLAG_FILL
        elif line.isupper() and line.strip():
            cell.font = BOLD_CELL
        else:
            cell.font = CELL_FONT
        r += 1
    ws.column_dimensions["A"].width = 100


if __name__ == "__main__":
    results_path = os.path.join(utils.TABLES_DIR, "knowledge_score_results.csv")
    df = pd.read_csv(results_path)

    out_path = os.path.join(utils.MODELS_DIR, "Knowledge_Score_Underserved_18to24_Report.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="_placeholder")
        write_results_sheet(writer, df)
        build_methodology_sheet(writer)
        wb = writer.book
        wb.move_sheet("Methodology", offset=-(len(wb.sheetnames) - 1))
        del wb["_placeholder"]

    print(f"Saved: {out_path}")
